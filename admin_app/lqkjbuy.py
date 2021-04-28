from django.shortcuts import HttpResponse
from django.db import connection
import json
from admin_app import public
from admin_app import models
import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment,Border,Side
import os
filename=os.path.basename(__file__)
from decimal import Decimal
import pymssql
import calendar
import sys

def main(request):
    log = public.logger

    log.info('----------------------%s-begin---------------------------'%filename)
    if request.method == "POST": #POST请求
        #请求body转为json
        tmp=request.body
        tmp=tmp.decode(encoding='utf-8')
        request_body=json.loads(tmp)
        trantype=request_body['trantype']
        log.info('trantype=[%s]' % trantype)
        fun_name = trantype
        if globals().get(fun_name):
            resp = globals()[fun_name](request, request_body)
        else:
            s = public.setrespinfo({"respcode": "100000", "respmsg": "api error! lqkjerp!"})
            resp = HttpResponse(s)
    elif request.method == "GET": #GET请求
        s = public.setrespinfo({"respcode": "100000", "respmsg": "api error"})
        resp = HttpResponse(s)
    log.info('----------------------%s-end---------------------------'%filename)
    return resp

#函数装饰器
def func_wrapper(func):
    def wrapper(request, request_body,*args,**kwargs):
        log = public.logger
        fun_name=func.__name__
        log.info('----------------------%s-%s-begin---------------------------'%(filename,fun_name))
        jsondata = {
            "respcode": "000000",
            "respmsg": "交易成功",
            "trantype": request_body.get('trantype', None),
        }
        res=func(jsondata,log,request, request_body,*args,**kwargs)
        s = json.dumps(jsondata, cls=models.JsonCustomEncoder, ensure_ascii=False)
        log.info(s)
        log.info('----------------------%s-%s-end---------------------------'%(filename,fun_name))
        return HttpResponse(s)
    return wrapper


############################################3


#######################################################################
number_keys=[
            'buy_amount',
            'aready_pay_amount',
            'need_pay_amount',
            'pay_amount_end_time',
            'need_pay_amount_not_end_time'
        ]

def _formatDate(date):
    # fmt = '%Y-%m-%d %H:%M:%S'
    fmt = '%Y-%m-%d'
    time_tuple = time.strptime(date, fmt)
    format_date = datetime.datetime(*time_tuple[:6])
    return format_date

#获取到期需付款按月分组
def get_need_pay_group_by_month(log,cursor,cre_id,already_pay,ass_period,begin_date, end_date):
    """
    :param log: 
    :param cursor: 
    :param cre_id: 信用代码
    :param already_pay: 已付款
    :param ass_period: 账期
    :return: 
    """
    # sql = "select code_amo_money,code_date,code_num from yw_payable_invoice_info where code_cre_id=%s and CODE='1' " \
    #       "and date_add(code_date,INTERVAL %s DAY) between %s and %s order by code_date"
    sql = "select code_amo_money,code_date,code_num from yw_payable_invoice_info where code_cre_id=%s and CODE='1' " \
          "and date_add(code_date,INTERVAL %s DAY) <= %s order by code_date"
    cursor.execute(sql, (cre_id,ass_period, end_date))
    rows = cursor.fetchall()
    kv={}
    code_detail=[]

    if rows:
        for code_amo_money,code_date,code_num in rows:
            already_pay-=code_amo_money
            if already_pay<0:
                code_date=code_date+datetime.timedelta(days=ass_period)
                key='%s年%s月'%(code_date.strftime('%Y'),code_date.strftime('%m'))
                if already_pay<0 and already_pay+code_amo_money>0:#只有一半的情况
                    money=-already_pay
                else:
                    money=code_amo_money
                if kv.get(key)==None:
                    kv[key]={}
                if kv[key]:
                    kv[key] += round(money,6)
                else:
                    kv[key ]= round(money,6)
                #统计未付发票详情
                temp={}
                temp['code_num']=code_num
                temp['code_date']=code_date.strftime('%Y-%m-%d %H:%M:%S')
                temp['pay_already']='%.2f'% (code_amo_money-money)
                temp['pay_not']='%.2f'% money
                code_detail.append(temp)
    else:
        return {},{}

    return kv,code_detail


@func_wrapper
def get_excel_data(jsondata,log,request, request_body):
    begin_date = request_body.get('begin_date')
    end_date = request_body.get('end_date')
    begin_cus = request_body.get('begin_cus')
    end_cus = request_body.get('end_cus')
    if not begin_date or  not end_date:
        jsondata['respcode'] = '400000'
        jsondata['respmsg'] = '开始日期和结束日期不可为空'
        return

    begin_date = _formatDate(begin_date)
    end_date = _formatDate(end_date)

    #获取所选日期的当月第一天和最后一天
    # now = end_date
    now = datetime.datetime.now() #取当前日期的
    year = now.year
    month = now.month
    today = calendar.monthrange(year, month)[1]  ## 最后一天
    month_start_day = datetime.date(year, month, 1)
    month_end_day = datetime.date(year, month, today)

    # 获取上月月份ID, 上月的第一天和最后一天
    last_day = month_start_day + datetime.timedelta(days=-1)
    last_month_start_day = datetime.date(last_day.year, last_day.month, 1)  # 当月第一最后一天
    month_last_day = calendar.monthrange(last_day.year, last_day.month)[1]  ## 最后一天的天数
    last_month_end_day = datetime.date(last_day.year, last_day.month, month_last_day)  # 当月最后一天

    # try:
    #     sqlserver_conn = pymssql.connect(server='10.10.1.250', user='sa', password='luyao123KEJI',
    #                                      database="db_18",
    #                                      timeout=20, autocommit=False)  # 获取连接  #sqlserver数据库链接句柄
    #     sqlserver_cur = sqlserver_conn.cursor()  # 获取光标
    # except Exception as ex:
    #     log.info(str(ex))
    #     jsondata['respcode'] = '999999'
    #     jsondata['respmsg'] = '连接ERP数据库错误'
    #     return

    cursor = connection.cursor()
    if begin_cus and end_cus:#查询指定客户信息
        sql="select fac_code,fac_rank,fac_name,cre_id,ass_period from yw_payable_factory_info where fac_code>=%s and fac_code<=%s and state='1'"
        # log.info(sql%(begin_cus,end_cus))
        cursor.execute(sql,(begin_cus,end_cus))
    else:#查询所有客户信息
        sql = "select fac_code,fac_rank,fac_name,cre_id,ass_period from yw_payable_factory_info where state='1'"
        # log.info(sql)
        cursor.execute(sql)
    facs = cursor.fetchall()
    detail=[]
    for fac_code,fac_rank,fac_name,cre_id,ass_period in facs:
        fac={}
        fac['fac_code']=fac_code
        fac['fac_rank']=fac_rank
        fac['fac_name']=fac_name
        fac['cre_id']=cre_id
        fac['ass_period'] = ass_period

        #已入发票金额总计
        # sql="select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and CODE='1' " \
            # "and code_tran_date between '%s' and '%s'" % ( cre_id, begin_date, end_date)
        sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and CODE='1' " \
              "and code_tran_date <= '%s' " % (cre_id, end_date)
        # log.info(sql)
        cursor.execute(sql)
        row=cursor.fetchone()
        if row:
            fac['buy_amount']='%.2f'% row[0]
        else:
            fac['buy_amount'] = '%.2f'% 0

        #已付款总计
        # sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and CODE='2' " \
        #       "and code_tran_date between '%s' and '%s'" % ( cre_id, begin_date, end_date)
        # sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and CODE='2' " \
        #       % (cre_id )
        sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and CODE='2' " \
              "and code_tran_date < '%s'" % (cre_id, end_date)
        # log.info(sql)
        cursor.execute(sql)
        row = cursor.fetchone()
        if row:
            fac['aready_pay_amount'] = '%.2f'% row[0]
        else:
            fac['aready_pay_amount'] = '%.2f'% 0

        #应付款总计
        fac['need_pay_amount']='%.2f' % (float(fac['buy_amount'])-float(fac['aready_pay_amount']) )

        #到期应付款
        ##到期采购总额
        sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id=%s and `CODE`='1' and date_add(code_date,INTERVAL %s DAY)<=%s"
        # log.info(sql % (cre_id,ass_period,end_date))
        cursor.execute(sql, (cre_id,ass_period,end_date))
        row = cursor.fetchone()
        if row:
            buy_amount_end_time = '%.2f'% row[0]
        else:
            buy_amount_end_time = '%.2f'% 0
        ##到期应付款=到期采购金额-已付款
        fac['pay_amount_end_time']='%.2f'% (float(buy_amount_end_time)-float(fac['aready_pay_amount']) )

        #到期额按时间分类
        fac['pay_amount_group_by_month']=''
        kv,code_datail=get_need_pay_group_by_month(log,cursor,cre_id,Decimal(fac['aready_pay_amount']),ass_period, begin_date, end_date)
        for key,value in kv.items():
            fac['pay_amount_group_by_month']+='%s:%.2f\n' % ( key, value )
        # log.info(str(fac['pay_amount_group_by_month']) )
        # log.info(str(type(fac['pay_amount_group_by_month'])))
        fac['pay_amount_group_by_month']= fac['pay_amount_group_by_month'][:-1]

        #未到期应付款
        sql = "select IFNULL(sum(code_amo_money),0) from yw_payable_invoice_info where code_cre_id='%s' and `CODE`='1' " \
              "and date_add(code_date,INTERVAL %s DAY)  > '%s'" % (cre_id, ass_period, end_date)
        # log.info(sql % (cre_id, ass_period, end_date))
        cursor.execute(sql )
        row = cursor.fetchone()
        if row:
            fac['need_pay_amount_not_end_time'] = '%.2f'% row[0]
        else:
            fac['need_pay_amount_not_end_time'] = '%.2f'% 0

        fac['code_datail']= code_datail

        #add by litz, 增加从erp中获取的两列
        # 暂估未进货明细
        # sql = "select ISNULL(sum(t.AMT), 0) from TF_ZG t, MF_ZG h where t.zg_id='ZG' and  t.zg_no=h.zg_no and not exists " \
        #       "( select 1 from TF_PSS where OS_NO=t.zg_no and PRD_NO=t.prd_no ) and h.cus_no='%s' " \
        #       "and CONVERT(varchar(100), h.zg_dd, 120) < '%s.000' " % (fac_code, end_date)
        # sql = "SELECT sum(B.AMT) FROM MF_ZG H " \
        #       "LEFT OUTER JOIN SALM S ON S.SAL_NO = H.SAL_NO " \
        #       "LEFT OUTER JOIN CASN N ON N.CAS_NO = H.CAS_NO LEFT OUTER JOIN TASK K ON H.CAS_NO = K.CAS_NO AND K.TASK_ID = H.TASK_ID " \
        #       "LEFT OUTER JOIN DEPT D ON D.DEP = H.DEP, TF_ZG B " \
        #       "LEFT OUTER JOIN MY_WH M ON M.WH = B.WH " \
        #       "LEFT OUTER JOIN TF_POS T ON (T.OS_ID = B.OS_ID) AND (T.OS_NO = B.OS_NO) AND (T.EST_ITM = B.EST_ITM) " \
        #       "LEFT OUTER JOIN MF_POS N1 ON (N1.OS_ID = B.OS_ID) AND (N1.OS_NO = B.OS_NO) AND (B.OS_ID = 'PO'), CUST C " \
        #       "LEFT OUTER JOIN AREA R ON R.AREA_NO = C.CUS_ARE, PRDT P " \
        #       "LEFT OUTER JOIN INDX Y ON Y.IDX_NO = P.IDX1 " \
        #       "WHERE (H.ZG_ID = 'ZG') AND (H.ZG_ID = B.ZG_ID) AND (H.ZG_NO = B.ZG_NO) AND (H.CUS_NO = C.CUS_NO) " \
        #       "AND (B.PRD_NO = P.PRD_NO) AND (H.ZG_DD <= '%s') AND (H.CUS_NO= '%s')  " \
        #       "AND (H.CLS_ID = 'F') AND ( IsNull(H.Cancel_ID, '') <> 'T' )" % (end_date, fac_code)
        # sql = "select sum(num*up) from yw_payable_zg_wjh where zg_dd between '%s' and '%s' and  cust_no='%s' " \
        #       % (begin_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), fac_code)
        sql = "select sum(num*up) from yw_payable_zg_wjh where zg_dd <= '%s' and  cust_no='%s' " \
              % ( end_date.strftime('%Y-%m-%d'), fac_code)
        # log.info( "暂估未进货明细:"+sql )
        cursor.execute(sql )
        row = cursor.fetchone()
        if row and row[0]:
            fac['buy_unbill_amt'] = '%.2f' % row[0]
        else:
            fac['buy_unbill_amt'] = '%.2f'% 0

        # 暂估入库明细，有质检却还没有进货单
        # sql = "select ISNULL(sum(t.up*b.qty), 0) from MF_TI h, TF_TI b, TF_POS t where h.TI_NO=b.TI_NO and b.BIL_NO=t.OS_NO " \
        #       "and b.prd_no=t.prd_no and b.est_itm=t.itm and h.CUS_NO='%s' " \
        #       "and CONVERT(varchar(100), h.ti_dd, 120) between '%s 00:00:00' and '%s 23:59:59' " % (fac_code, month_start_day, month_end_day)
        # sql = "SELECT  sum(B.AMT) FROM MF_ZG A Left Outer Join CUST C On C.CUS_NO=A.CUS_NO  " \
        #       "Left Outer Join TASK Q On Q.CAS_NO=A.CAS_NO AND Q.TASK_ID=A.TASK_ID " \
        #       "Left Outer Join CASN R On R.CAS_NO=A.CAS_NO " \
        #       "Left Outer Join DEPT D ON D.DEP=A.DEP " \
        #       "Left Outer Join SALM S On S.SAL_NO=A.SAL_NO,TF_ZG B  " \
        #       "Left Outer Join MY_WH M On M.WH=B.WH " \
        #       "Left Outer Join PRDT E On E.PRD_NO=B.PRD_NO " \
        #       "Left Outer Join BAT_NO TT On TT.BAT_NO=B.BAT_NO " \
        #       "Left Outer Join TF_POS T On (T.OS_ID=B.OS_ID) And (T.OS_NO=B.OS_NO) and (T.EST_ITM=B.EST_ITM) " \
        #       "Left Outer Join MF_POS N On (N.OS_ID=B.OS_ID)And(N.OS_NO=B.OS_NO)And(B.OS_ID='PO') " \
        #       "Left Outer Join INDX H On H.IDX_NO=E.IDX1 " \
        #       "Where (B.ZG_ID='ZG') AND (B.ZG_ID=A.ZG_ID) AND (B.ZG_NO=A.ZG_NO) AND (B.PRD_NO=E.PRD_NO) " \
        #       "And (A.ZG_DD >='%s') And (A.ZG_DD <='%s') And (A.CUS_NO='%s') " \
        #       "And (IsNull(A.Cancel_ID,'')<>'T') " % (month_start_day, month_end_day, fac_code)
        sql = "select sum(amt) from yw_payable_zg_rk where zg_dd between '%s' and '%s' and  cust_no='%s' " \
              % (last_month_start_day.strftime('%Y-%m-%d'), last_month_end_day.strftime('%Y-%m-%d'), fac_code) #上月入库金额
        log.info( "暂估入库明细:"+sql )
        cursor.execute(sql )
        row = cursor.fetchone()
        if row and row[0]:
            fac['lastmonth_insamt'] = '%.2f' % row[0]
        else:
            fac['lastmonth_insamt'] = '%.2f' % 0

        #获取上月开票金额
        sql =  "select sum(code_amo_money) from yw_payable_invoice_info a, yw_payable_factory_info b " \
               "where a.code_cre_id=b.cre_id and a.code='1' and a.code_tran_date between '%s' and '%s' " \
               "and b.fac_code='%s' " % (last_month_start_day, last_month_end_day, fac_code)
        cursor.execute(sql)
        row = cursor.fetchone()
        if row and row[0]:
            fac['lastmonth_billamt'] = '%.2f' % row[0]
        else:
            fac['lastmonth_billamt'] = '%.2f' % 0

        # 判断如果数值全0不统计
        flag = False
        for key, value in fac.items():
            if key in number_keys:
                # 如果有不等于0的 flag=True
                if float(value) != float(0):
                    flag = True
                    # 保留6位小数
                    fac[key] = value
        # 如果全0 跳过
        # log.info('flag=%s' % flag)
        if not flag:
            continue

        detail.append(fac)

    jsondata['tableData']=detail
    jsondata['tableHead']=[
        {
            'name': 'fac_code',
            'label': '供应商代码',
            'width': 150
        },
        {
            'name': 'fac_name',
            'label': '厂商名称',
            'width': 150
        },
        {
            'name': 'ass_period',
            'label': '账期(天)',
            'width': 150
        },
        {
            'name': 'buy_amount',
            'label': '已入发票金额',
            'width': 150
        },
        {
            'name': 'aready_pay_amount',
            'label': '已付款总计',
            'width': 150
        },
        {
            'name': 'buy_unbill_amt',
            'label': '采购未开票',
            'width': 150
        },
        {
            'name': 'lastmonth_insamt',
            'label': '上月采购入库金额',
            'width': 150
        },
        {
            'name': 'lastmonth_billamt',
            'label': '上月开票金额',
            'width': 150
        },
        {
            'name': 'need_pay_amount',
            'label': '应付款总计',
            'width': 150
        },
        {
            'name': 'pay_amount_end_time',
            'label': '到期应付款',
            'width': 150
        },
        {
            'name': 'pay_amount_group_by_month',
            'label': '到期额按时间分类',
            'width': 200
        },
        {
            'name': 'need_pay_amount_not_end_time',
            'label': '未到期应付款',
            'width': 150
        },


    ]

    if begin_cus and end_cus:
        filename = '联桥科技应付账款_%s_%s-%s_%s.xlsx' % (begin_cus,end_cus, begin_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d'))
    else:
        filename='联桥科技应付账款-%s_%s.xlsx'%(begin_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d'))
    _make_excel(jsondata['tableHead'], jsondata['tableData'], filename, log, begin_date, end_date)
    jsondata['fileurl']='http://10.10.10.174/static/filedown/%s'%filename


def _getGbkLen(w):
    if type(w)==str and w.find('\n')!=-1:
        w=w.split('\n')[0]
    return len(str(w).encode('gbk'))

def _make_excel(tableHead,tableData,filename,log, begin_date, end_date):
    # 拼装数据
    fullData = []
    temp=[]
    temp.append('序号')
    for j in range(len(tableData)+1):
        temp.append(str(j+1))
    fullData.append(temp)
    for i in range(len(tableHead)):
        temp = []
        label = tableHead[i]['label']
        name = tableHead[i]['name']
        temp.append(label)
        for j in range(len(tableData)):
            temp.append(tableData[j][name])
        if name == 'cus_code':
            temp.append('总计：')
        elif name in number_keys:
            b = 0.00
            for itm in temp[1:]:
                b = b + float(itm)
            temp.append('%.2f' % b)
        else:
            temp.append('')
        fullData.append(temp)
    null_field=['本月申请','承兑','备注']
    for label in null_field:
        temp=[]
        temp.append(label)
        for j in range(len(tableData)+1):
            temp.append('')
        fullData.append(temp)

    wb = Workbook()
    mySheet = wb.create_sheet(index=0, title="Mysheet")

    fontStyle = Font(size=18)
    fontStyle2 = Font(size=14)
    fontMini = Font(size=10)
    centerStyle = Alignment(horizontal='center', vertical='center',wrap_text=True)
    borderStyle = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')

    )

    startLetter = chr(65)
    endLetter = chr(65 + len(fullData) - 1)
    mySheet.merge_cells('%s1:%s1' % (startLetter, endLetter))
    mySheet.merge_cells('%s2:%s2' % (startLetter, endLetter))
    mySheet.cell(1, 1).value = '联桥科技有限公司'
    mySheet.cell(1, 1).alignment = centerStyle
    mySheet.cell(1, 1).font = fontStyle
    mySheet.cell(2, 1).value = '%s年%s月份应付账款报表'%(end_date.strftime('%Y'),end_date.strftime('%m'))
    mySheet.cell(2, 1).alignment = centerStyle
    mySheet.cell(2, 1).font = fontStyle2

    #日期段
    mySheet.cell(3, 5).value = '开始日期:'
    mySheet.cell(3, 6).value = '%s' % (begin_date.strftime('%Y-%m-%d'))
    mySheet.cell(3, 7).value = '截止日期:'
    mySheet.cell(3, 8).value = '%s' % (end_date.strftime('%Y-%m-%d'))
    for i in range(5, 9):
        mySheet.cell(3, i).alignment = centerStyle
        mySheet.cell(3, i).font = fontMini

    # mySheet.column_dimensions[chr(65 + (len(fullData) - 2) - 1)].width=100

    row = 0
    column = 1
    max_line_num = {}
    for items in fullData:
        row = 4
        tempSorted = sorted(items, key=lambda x: -_getGbkLen(x))
        maxlen = _getGbkLen(tempSorted[0]) + 4

        # 调整列高
        mySheet.column_dimensions[chr(65 + column - 1)].width = maxlen
        for item in items:
            # 调整行高
            if type(item)==str:
                line_num = len(item.split('\n'))  # 行数
            else:
                line_num=1
            if line_num>max_line_num.get(str(row),0):
                max_line_num[str(row)]=line_num
            # mySheet.row_dimensions[row].height = 25*line_num
            mySheet.cell(row, column).value = item
            mySheet.cell(row, column).alignment = centerStyle
            mySheet.cell(row, column).border = borderStyle
            row += 1
        column += 1



    row += 1
    column = 2
    mySheet.cell(row, column).value = '编制：'
    mySheet.cell(row, column).alignment = centerStyle

    column = int(len(tableHead) / 2)
    mySheet.cell(row, column).value = '审核：'
    mySheet.cell(row, column).alignment = centerStyle

    column = len(tableHead) - 2
    mySheet.cell(row, column).value = '审批：'
    mySheet.cell(row, column).alignment = centerStyle

    # 调整行高
    mySheet.row_dimensions[1].height=30
    mySheet.row_dimensions[2].height = 25
    for row,line_num in max_line_num.items():
        mySheet.row_dimensions[int(row)].height = 22 * line_num


    wb.save(public.localhome+'static/filedown/%s'%filename)
    wb.close()

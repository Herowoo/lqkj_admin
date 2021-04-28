#coding:utf-8
# -*- coding: utf-8 -*-
import hashlib

from django.shortcuts import  HttpResponse
from django.db import connection
import json
from admin_app import public, userrole
from admin_app import models
import datetime
import os
import base64
import openpyxl
import zipfile
from admin_app.TempRender2 import TempRender

#文件处理
def admin(request):
    log = public.logger
    log.info('----------------------filecfg_begin---------------------------')

    if request.method == "POST":
        # #请求body转为json
        tmp=request.body
        tmp=tmp.decode(encoding='utf-8')
        reqest_body=json.loads(tmp)

        trantype=reqest_body['trantype']
        # print('-'*20,trantype,'-'*20)
        log.info('trantype=[%s]' % trantype)

        if trantype == 'upload':  #文件二进制流的方式上传文件
            if public.checksession(request) == False:
                s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
                respinfo = HttpResponse(s)
                return respinfo
            resp = fileupload(request, reqest_body)
        elif trantype == 'files_upload':  #
            # if public.checksession(request) == False:
            #     s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
            #     respinfo = HttpResponse(s)
            #     return respinfo
            resp = files_upload(request, reqest_body)
        elif trantype == 'files_download':  #
            # if public.checksession(request) == False:
            #     s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
            #     respinfo = HttpResponse(s)
            #     return respinfo
            resp = files_download(request, reqest_body)
        elif trantype == 'export2excel':  #导出数据到excel文件
            # if public.checksession(request) == False:
            #     s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
            #     respinfo = HttpResponse(s)
            #     return respinfo
            resp = export2excel(request, reqest_body)
        elif trantype == 'imageupload':  #文件二进制流的方式上传文件
            if public.checksession(request) == False:
                s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
                respinfo = HttpResponse(s)
                return respinfo
            resp = imageupload(request, reqest_body)
        elif trantype == 'componentupload':  #文件二进制流的方式上传多个文件,前端敏捷组件上传
            if public.checksession(request) == False:
                s = public.setrespinfo({"respcode": "100001", "respmsg": "请重新登陆"})
                respinfo = HttpResponse(s)
                return respinfo
            resp = componentupload(request, reqest_body)

        elif trantype == 'GetComponentList':  #获取组件列表,前端敏捷开发
            resp = GetComponentList(request, reqest_body)

        else:
            s = public.setrespinfo({"respcode": "100000", "respmsg": "api error"})
            resp = HttpResponse(s)

    elif request.method == "GET":
        s = public.setrespinfo({"respcode": "000000", "respmsg": "api error"})
        resp = HttpResponse(s)

    log.info('----------------------filecfg_end---------------------------')
    return resp

#文件资源上传，反回md5值的url服务器路径
def files_upload(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-files_upload-begin---------------------------')
    filename=reqest_body.get('filename',None)
    if filename==None or len(filename)<2:
        s = public.setrespinfo({"respcode": "323311", "respmsg": "上送文件名错误"})
        return s
    uid=reqest_body.get('uid',None)
    file = reqest_body.get('file', None)
    if file==None:
        s = public.setrespinfo({"respcode": "323312", "respmsg": "上送文件内容错误"})
        return s

    filename_ext=filename.split('.')[1]
    #生成md5值的文件名
    m2=hashlib.md5()
    m2.update(file.encode('raw_unicode_escape'))
    md5_filename=m2.hexdigest()+'.'+filename_ext

    # 保存文件到本地文件上传目录
    filepath=public.localhome+'fileup/'
    file_name = open(filepath+md5_filename, 'wb')
    mylen = len(reqest_body['file'])
    log.info('mylen=' + str(mylen))
    file_name.write(file.encode('raw_unicode_escape')) #前端在json报文中，把二进制当字符串上送了，可以这样转换
    file_name.close()

    #插入数据库
    cur = connection.cursor()
    sql = "insert into yw_workflow_file(flow_id,file_name,md5_name,state,tran_date,user_id) value(%s,%s,%s,%s,%s,%s)"
    res=cur.execute(sql, (1,filename,md5_filename,'1',datetime.datetime.now(),uid))
    fileid=cur.lastrowid
    cur.close()

    data = {
        "respcode": "000000",
        "respmsg": "上传成功",
        "trantype": reqest_body.get('trantype', None),
        "fileid":str(fileid),
        'filename':filename
    }

    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-files_upload-end---------------------------')
    return HttpResponse(s)



#文件资源下载
def files_download(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-files_upload-begin---------------------------')
    fileid=reqest_body.get('fileid',None)
    if fileid==None:
        s = public.setrespinfo({"respcode": "323311", "respmsg": "资源id有误"})
        return s

    cur = connection.cursor()
    sql = "select * from yw_workflow_file where id=%s"
    cur.execute(sql, fileid)
    rows = cur.fetchone()
    if not rows:
        s = public.setrespinfo({"respcode": "323311", "respmsg": "资源id有误"})
        return s
    print('rows=',rows)
    filename= rows[2]
    md5_filename = rows[3]
    cur.close()


    # 本地文件上传目录
    filepath=public.localhome+'fileup/'
    file = open(filepath+md5_filename, 'rb')
    # response = HttpResponse(file)

    data = {
        "respcode": "000000",
        "trantype": reqest_body.get('trantype', None),
        "data": base64.b64encode(file.read()).decode(),
        "filename":filename
    }

    s = json.dumps(data)
    # response['Content-Type'] = 'application/octet-stream'
    # response['Content-Disposition'] = 'attachment;filename="%s"'%filename
    return HttpResponse(s)



# data = {
#         "respcode": "000000",
#         "respmsg": "上传成功",
#         "trantype": reqest_body.get('trantype', None),
#         "file":file
#     }
#     file.close()
#
#     s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
#     log.info('----------------------Admin-files_upload-end---------------------------')
#     return HttpResponse(s)


#excel文件上传,数据导入
def fileupload(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-fileupload-begin---------------------------')
    filename=reqest_body.get('filename',None)
    if filename==None or len(filename)<2:
        s = public.setrespinfo({"respcode": "323311", "respmsg": "上送文件名错误"})
        return s

    file = reqest_body.get('file', None)
    if file==None:
        s = public.setrespinfo({"respcode": "323312", "respmsg": "上送文件内容错误"})
        return s

    menuid = reqest_body.get('MENU_ID', None)
    if menuid==None:
        s = public.setrespinfo({"respcode": "323310", "respmsg": "上送菜单编号错误"})
        return s

    #保存文件到本地文件上传目录，如果有文件名重复，则覆盖原文件
    filepath=public.localhome+'fileup/'
    file_name = open(filepath+filename, 'wb')
    mylen = len(reqest_body['file'])
    log.info('mylen=' + str(mylen))
    file_name.write(file.encode('raw_unicode_escape')) #前端在json报文中，把二进制当字符串上送了，可以这样转换
    file_name.close()

    file_name = filepath + filename
    file_rows=int( getexcellinenum( file_name ) ) -1 #文件行数，-1是去掉excel文件中的标题

    uid = reqest_body.get('uid', None)
    if file==None:
        s = public.setrespinfo({"respcode": "323313", "respmsg": "操作用户字段上送有误"})
        return s
    # 查找管理台用户信息表权限
    log.info('user_id=' + str(uid))
    try:
        AdminUser = models.IrsadminUser.objects.get(user_id=uid)
    except models.IrsadminUser.DoesNotExist:
        s = public.setrespinfo({"respcode": "323314", "respmsg": "操作用户不存在"})
        return s

    #插入文件上传登记表一条记录
    FileInfo=models.IrsadminUnfileInfo(
        user_id = AdminUser.user_id,
        ip = request.META.get('HTTP_X_FORWARDED_FOR',None),
        tran_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        menu_id = menuid,
        file_name = filename,
        file_rows = file_rows,
        deal_num=0,
        resp_code= '888888',
        resp_msg = '正在处理中',
    )
    FileInfo.save()

    log.info('-----开始处理文件-----')
    flag,resp=Excel_WriteDB(FileInfo.file_id, menuid, file_name)  # 子进程处理文件入库的操作
    if flag == True:
        data = {
            "respcode": "000000",
            "respmsg": "批量数据导入成功",
            "trantype": reqest_body.get('trantype', None),
            "file_id":FileInfo.file_id
        }
    else:
        data = resp

    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-fileupload-end---------------------------')
    return HttpResponse(s)

#压缩指定文件夹
def zipDir(dirpath, outFullName):
        """
        压缩指定文件夹
        :param dirpath: 目标文件夹路径
        :param outFullName: 压缩文件保存路径+xxxx.zip
        :return: 无
        """
        zip = zipfile.ZipFile(outFullName, "w", zipfile.ZIP_DEFLATED)
        for path, dirnames, filenames in os.walk(dirpath):
            # 去掉目标跟路径，只对目标文件夹下边的文件及文件夹进行压缩
            fpath = path.replace(dirpath, '')

            for filename in filenames:
                zip.write(os.path.join(path, filename), os.path.join(fpath, filename))
        zip.close()

#多个文件上传，前端敏捷开发组件上传
def componentupload(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-componentupload-begin---------------------------')

    filelist=reqest_body.get('filelist',None)
    imagelist = reqest_body.get('imagelist', None)  #图片列表
    if filelist==None or len(filelist)==0:
        if imagelist == None or len(imagelist) == 0:
            s = public.setrespinfo({"respcode": "323411", "respmsg": "上送文件列表为空"})
            return s

    menuid = reqest_body.get('MENU_ID', None)
    if menuid==None:
        s = public.setrespinfo({"respcode": "323310", "respmsg": "上送菜单编号错误"})
        return s

    #判断用户角色权限
    #校验权限
    flag=userrole.checkauth(request.session.get('user_id', None), menuid)
    if flag==False:
        s = public.setrespinfo({"respcode": "349921", "respmsg": "用户权限不正确!"})
        return HttpResponse(s)

    #判断文件夹是否存在，如果不存在，则新建
    component_name = reqest_body.get('component_name', None)
    if component_name==None or component_name=='':
        s = public.setrespinfo({"respcode": "323412", "respmsg": "组件名不可为空"})
        return s
    component_tag = reqest_body.get('component_tag', None)
    if component_tag == None or component_tag=='':
        s = public.setrespinfo({"respcode": "323413", "respmsg": "组件标签不可为空"})
        return s

    #组件绝对路径
    filepath = '/home/irsqf/web-library/src/components/Library/'+component_name
    folder = os.path.exists(filepath)
    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(filepath)  # makedirs 创建文件时如果路径不存在会创建这个路径
    folder = os.path.exists(filepath+'/images')  #图片文件夹
    if not folder:
        os.makedirs(filepath+'/images')

    dealtype = reqest_body.get('dealtype', None)#处理方式  1:新增 2:覆盖

    #判断文件是否已经存在
    if dealtype != 2:
        for item in filelist:
            file = filepath + '/' + item['filename']
            if os.path.exists(file): #文件是否已经存在？？
                s = public.setrespinfo({"respcode": "323499", "respmsg": "文件已经存在，请选择覆盖方式!"})
                return s
        for item in imagelist:
            file = filepath + '/images/' + item['filename']
            if os.path.exists(file):  # 文件是否已经存在？？
                s = public.setrespinfo({"respcode": "323499", "respmsg": "文件已经存在，请选择覆盖方式!"})
                return s

    #保存文件到指定文件上传目录，
    for item in filelist:
        file = filepath + '/' + item['filename']
        file_name = open(file, 'wb')
        file_name.write(item['file'].encode('raw_unicode_escape')) #前端在json报文中，把二进制当字符串上送了，可以这样转换
        file_name.close()
    #保存图片文件到指定目录
    for item in imagelist:
        file = filepath + '/images/' + item['filename']
        log.info(file)
        file_name = open(file, 'wb')
        file_name.write(item['file'].encode('raw_unicode_escape')) #前端在json报文中，把二进制当字符串上送了，可以这样转换
        file_name.close()


    #生成zip文件方便下载时使用
    ZipFileName = '/home/irsqf/web-library/src/download/' + component_name + '.zip'
    zipDir(filepath, ZipFileName)

    #判断是否有记录
    try:
        ComponentInfo = models.IrsServerVueComponents.objects.filter(comp_path=component_name)
    except models.IrsServerVueComponents.DoesNotExist:
        ComponentInfo = None

    if ComponentInfo.__len__()==0:
        # 插入组件登记表一条记录
        ComponentInfo = models.IrsServerVueComponents(
            user_id=request.session.get('user_id', None),
            create_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            comp_name=reqest_body.get('component_menuname', reqest_body.get('component_tag', None)),
            comp_path=component_name,
            comp_tag=reqest_body.get('component_tag', None),
            status='1',
        )
        ComponentInfo.save()
    else:
        #更新记录
        ComponentInfo.update(
            user_id=request.session.get('user_id', None),
            update_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            comp_name=reqest_body.get('component_menuname', reqest_body.get('component_tag', None)),
            comp_path=component_name,
            comp_tag=reqest_body.get('component_tag', None),
            status='1',
        )

    data = {
        "respcode": "000000",
        "respmsg": "文件上传成功",
        "trantype": reqest_body.get('trantype', None),
    }

    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-componentupload-end---------------------------')
    return HttpResponse(s)

#图片资源上传，反回md5值的url服务器路径
def imageupload(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-imageupload-begin---------------------------')
    filename=reqest_body.get('filename',None)
    if filename==None or len(filename)<2:
        s = public.setrespinfo({"respcode": "323311", "respmsg": "上送文件名错误"})
        return s

    file = reqest_body.get('file', None)
    if file==None:
        s = public.setrespinfo({"respcode": "323312", "respmsg": "上送文件内容错误"})
        return s

    filename_ext=filename.split('.')[1]
    #生成md5值的文件名
    m2=hashlib.md5()
    m2.update(file.encode('raw_unicode_escape'))
    filename=m2.hexdigest()+'.'+filename_ext

    # 保存文件到本地文件上传目录
    filepath=public.localhome+'static/images/detl/'
    file_name = open(filepath+filename, 'wb')
    mylen = len(reqest_body['file'])
    log.info('mylen=' + str(mylen))
    file_name.write(file.encode('raw_unicode_escape')) #前端在json报文中，把二进制当字符串上送了，可以这样转换
    file_name.close()

    data = {
        "respcode": "000000",
        "respmsg": "图片上传成功",
        "trantype": reqest_body.get('trantype', None),
        "url":request.META['HTTP_ORIGIN']+'/'+'static/images/detl/'+filename,
    }

    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-imageupload-end---------------------------')
    return HttpResponse(s)

#获取excel文件的行数
def getexcellinenum( file_name):
    try:
        wb = openpyxl.load_workbook(filename=file_name)
    except FileNotFoundError:
        print('文件不存在')
        return -1
    except:
        print('打开文件错误')
        return -2
    ws = wb.active
    linenum=ws.max_row-1 #行数 -1是标题
    wb.close()
    return linenum

#获取组件列表,前端敏捷开发组件浏览
def GetComponentList(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-readexcel-begin---------------------------')

    data = {
        "respcode": "000000",
        "respmsg": "交易成功",
        "trantype": reqest_body.get('trantype', None),
        "allTags": [],
        "detail":[],
    }

    # 判断是否有记录
    try:
        ComponentInfo = models.IrsServerVueComponents.objects.filter(status='1')
    except models.IrsServerVueComponents.DoesNotExist:
        ComponentInfo = None

    #组织返回报文
    for item in ComponentInfo:
        Components = {}
        Components['name']=item.comp_name
        Components['path'] = item.comp_path
        Components['info'] = item.notes
        Components['type'] = 'mobile'  #mobile, #pc
        data['detail'].append(Components)
        for tag in item.comp_tag.split(','):
            if tag == None or tag == '':
                continue
            if tag not in data['allTags']:
                data['allTags'].append(tag)



    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-imageupload-end---------------------------')
    return HttpResponse(s)

#将excel文件的内容写入数据库表中
def Excel_WriteDB(file_id, menuid, file_name ):
    log = public.logger
    log.info('----------------------Admin-readexcel-begin---------------------------')

    #根据 menuid 获取对应的表配置信息
    try:
        MenuTable = models.IrsadminMenu.objects.get(menu_id=menuid)
    except models.IrsadminMenu.DoesNotExist:
        MenuTable = None
        models.IrsadminUnfileInfo.objects.filter(file_id=file_id).update(
            resp_code='310002',
            resp_msg='菜单编号不存在',
        )
        data = {
            "respcode": "310002",
            "respmsg": "菜单编号不存在"
        }
        return False, data

    try:
        TranReg=models.IrsadminDbTranReg.objects.get(app_id=MenuTable.app_id)
    except models.IrsadminDbTranReg.DoesNotExist:
        TranReg = None
        models.IrsadminUnfileInfo.objects.filter(file_id=file_id).update(
            resp_code='310003',
            resp_msg='应用编号不存在',
        )
        data = {
            "respcode": "310003",
            "respmsg": "应用编号不存在"
        }
        return False, data

    try:
        TranList=models.IrsadminDbTranList.objects.filter(app_id=MenuTable.app_id, state='Y').order_by('order_id')
    except models.IrsadminDbTranReg.DoesNotExist:
        TranList = None
        models.IrsadminUnfileInfo.objects.filter(file_id=file_id).update(
            resp_code='310004',
            resp_msg='应用配置不存在',
        )
        data = {
            "respcode": "310004",
            "respmsg": "应用配置不存在"
        }
        return False, data

    fields=''
    values=''
    fields_num=0
    for item in TranList:
        fields_num = fields_num + 1
        if fields == '':
            fields = item.field_id
            values = values + '%s'
        else:
            fields = fields + ',' + item.field_id
            values = values + ',' + '%s'
    insertsql='insert into '+TranReg.table_name +'('+fields+') values('+ values + ')'
    log.info(insertsql)

    #读取excel数据并处理
    # print(file_name)
    try:
        wb = openpyxl.load_workbook(filename=file_name)
    except FileNotFoundError:
        data = {
            "respcode": "360001",
            "respmsg": "文件不存在"
        }
        return False, data
    except:
        log.error('打开文件错误', exc_info = True)
        data = {
            "respcode": "360002",
            "respmsg": "打开文件错误"
        }
        return False, data
    ws = wb.active
    # 从第一行获取, 获取标题,并判断模板是否正确
    i=0;
    for item in TranList:
        i=i+1
        # print(item.field_name)
        xl_cell = ws.cell(row=1, column=i).value
        if str(item.field_name) != str(xl_cell):
            wb.close()#内容有错误，关闭文件并返回错误
            data = {
                "respcode": "360003",
                "respmsg": "文件内容标题错误"+str(xl_cell)
            }
            return False, data

    #开始处理数据
    i=0
    for row in ws.rows:
        i=i+1
        if i==1: #第一行标题不处理
            continue
        values = []
        for col in row:
            log.info(col.value)
            if col.value == '':
                val=None
            else:
                val=col.value
            values.append(val)
        print(tuple(values))
        log.info(insertsql % tuple(values) )
        #执行写入数据库操作
        cur = connection.cursor()
        cur.execute(insertsql, tuple(values))
        cur.close()

    wb.close()

    log.info('----------------------Admin-readexcel-end---------------------------')
    return True, '文件入库成功'

#表数据写入excel文件
def export2excel(request, reqest_body):
    log = public.logger
    log.info('----------------------Admin-export2excel-begin---------------------------')

    menuTable=models.IrsadminMenu.objects.get(menu_id=reqest_body.get('MENU_ID',""))
    TranRegTable=models.IrsadminDbTranReg.objects.get(app_id=menuTable.app_id)

    #校验权限
    flag=userrole.checkauth(reqest_body.get('user_id', None), menuTable.menu_id)
    if flag==False:
        s = public.setrespinfo({"respcode": "349921", "respmsg": "用户权限不正确!"})
        return HttpResponse(s)

    pagesize = int(reqest_body.get('pagesize', "10"))
    pagenum = int(reqest_body.get('pagenum', "1"))

    searchinfo=reqest_body.get('searchinfo',None)
    whereinfo=""
    #print(type(searchinfo), searchinfo)
    if searchinfo:
        #有查询条件
        for item in searchinfo:
            selectid=item.get('selectid', None)
            selecttype=item.get('selecttype', None)
            selectvalue = item.get('selectvalue', None)

            if selectid==None or selectid=="":
                continue
            if selecttype==None or selecttype=="":
                continue

            # if item['uitype'] == 'datetime' and selectvalue:
            #     selectvalue = datetime.datetime.strptime(selectvalue, "%Y-%m-%d %H:%M:%S")
            if selecttype == 'like' and selectvalue:
                selectvalue="%"+selectvalue+"%"
            wheretemp=" and "+selectid+" "+selecttype+" '"+selectvalue+"'"
            #print(wheretemp)
            if whereinfo == "":
                whereinfo = wheretemp
            else:
                whereinfo = whereinfo+' '+wheretemp

    if TranRegTable.where_ctrl==None:
        TranRegTable.where_ctrl=""
    if TranRegTable.order_ctrl==None:
        TranRegTable.order_ctrl=""

    cur = connection.cursor()
    sql = "select FIELD_ID,FIELD_NAME from IRSADMIN_DB_TRAN_LIST where STATE='Y' and APP_ID=%s order by ORDER_ID asc"
    log.info(sql % menuTable.app_id)
    cur.execute(sql, menuTable.app_id)
    rows = cur.fetchall()
    fieldlist=""
    fieldnamelist = []
    for item in rows:
        if fieldlist.__len__()==0:
            fieldlist = item[0]
        else:
            fieldlist=fieldlist+','+item[0]
        fieldnamelist.append(item[1])

    if TranRegTable.where_ctrl==None or TranRegTable.where_ctrl=="":
        if whereinfo.__len__()>0:
            TranRegTable.where_ctrl="where 1=1"
    else:
        if 'WHERE' not in TranRegTable.where_ctrl.upper():
            TranRegTable.where_ctrl='where '+TranRegTable.where_ctrl
        if '${ORGID}' in TranRegTable.where_ctrl:
            orgidsql="(SELECT org_id FROM irsadmin_user_org WHERE user_id='%s')" % request.session.get('user_id', None)
            TranRegTable.where_ctrl = TranRegTable.where_ctrl.replace('${ORGID}', orgidsql)

    selsql = "select "+fieldlist+ " from "+ TranRegTable.table_name+" "+TranRegTable.where_ctrl+" "+whereinfo+" "+TranRegTable.order_ctrl
    log.info(selsql)
    cur.execute(selsql)
    rows = cur.fetchall()
    # totalnum=rows.__len__()
    fieldlist=fieldlist.split(',')
    #print("查询时获取的表头字段:",fieldlist)



    app_id=menuTable.app_id
    temp_path=public.localhome+'filetemplate/crudcfg_%s.xlsx'%(app_id)

    # 创建excel文件
    totalnum = 0
    file_name = TranRegTable.tran_desc + '_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx'
    if not os.path.exists(temp_path):#如果没有模板 直接写
        # print('file_name=',file_name)
        wb = openpyxl.Workbook()  # 创建文件对象
        ws = wb.active     #获取第一个sheet, 激活 worksheet
        try:
            #标题,第一行
            col=0
            for item in fieldnamelist:
                col = col + 1
                ws.cell(row=1, column=col).value = item

            totalnum=1  #标题已经占用了一行，从第二行开始
            for item in rows:
                totalnum=totalnum+1
                # print(item,item[0])
                col=0
                for subitem in item:
                    col = col + 1
                    ws.cell(row=totalnum, column=col).value = subitem
            wb.save( public.localhome+'static/filedown/'+file_name)  #保存excel文件
            wb.close()
            # fp=open(public.localhome+'static/filedown/'+file_name,'rb')
            # xlsbytes=fp.read()
            # fp.close()

            data = {
                "respcode": "000000",
                "respmsg": "交易成功",
                "trantype": reqest_body.get('trantype', None),
                "totalnum": totalnum - 1,
                "filename": file_name,
                # "file": xlsbytes.decode('raw_unicode_escape'),
                "fileurl": request.META['HTTP_ORIGIN'] + '/' + 'static/filedown/' + file_name
            }
        except:
            wb.close()
            log.error('导出数据到excel文件错误', exc_info=True)
            data = {
                "respcode": "000000",
                "respmsg": "导出数据到excel文件错误",
                "trantype": reqest_body.get('trantype', None)
            }
        finally:
            cur.close()
    else:#有模板 在模板基础上写
        try:
            tr = TempRender(temp_path,public.localhome+'static/filedown/'+file_name)
            #拼装数据
            kvs = {}
            for key in fieldlist:
                kvs[key]=[]
            for line in rows:
                for i,item in enumerate(line):
                    kvs[fieldlist[i]].append(item)
            #数据字典转换
            table_name=TranRegTable.table_name
            for key,value in kvs.items():
                dict_name=str(table_name).upper()+'.'+str(key).upper()
                cur = connection.cursor()
                sql = "select dict_code,dict_target from irs_ywty_dict where dict_name='%s'"
                log.info(sql % dict_name)
                cur.execute(sql%dict_name)
                rows = cur.fetchall()
                tran_dict = {}
                for row in rows:
                    tran_dict[str(row[0])]=row[1]
                log.info('【tran_dict=】%s' % tran_dict)
                #开始转换
                for index,item in enumerate(value):
                    new_value=tran_dict.get(str(item))
                    if not new_value:
                        new_value=item
                    kvs[key][index]=new_value
            log.info('【kvs=】%s'%kvs)


            temp = sorted(kvs.items(), key=lambda x: -len(x[1]))
            maxIndex=len(temp[0][1])
            kvs['index']=[i for i in range(1,maxIndex+1)]

            tr.render(kvs)
            tr.save()
            data = {
                "respcode": "000000",
                "respmsg": "交易成功",
                "trantype": reqest_body.get('trantype', None),
                "totalnum": totalnum - 1,
                "filename": file_name,
                # "file": xlsbytes.decode('raw_unicode_escape'),
                "fileurl": request.META['HTTP_ORIGIN'] + '/' + 'static/filedown/' + file_name
            }
        except Exception as e:
            log.error(e)
            # tr.close()
            log.error('导出数据到excel文件错误', exc_info=True)
            data = {
                "respcode": "000000",
                "respmsg": "导出数据到excel文件错误",
                "trantype": reqest_body.get('trantype', None)
            }
        finally:
            cur.close()

    s = json.dumps(data, cls=models.JsonCustomEncoder, ensure_ascii=False)
    log.info('----------------------Admin-export2excel-end---------------------------')
    return HttpResponse(s)


def write_excel():
    pass

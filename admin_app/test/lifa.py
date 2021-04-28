from functools import reduce
from zhdate import ZhDate
import datetime
import openpyxl
import itertools
from django.db import connection
import os
# from admin_app.sys import public_db
from pypinyin import lazy_pinyin


def readxlsx():
    filename = r'd:\测试excel1.xlsx'
    inwb = openpyxl.load_workbook(filename)
    sheetname1 = inwb.get_sheet_names()
    ws = inwb[sheetname1[0]]
    rang1 = ws['A1':'C2']
    return rang1


def dxweek(dt=datetime.date.today()):
    if dt.isocalendar()[1] % 2 == 0:
        print('大周')
    else:
        print('小周')


def not_null_response(respcode, **param):
    # res_msg = respcode + ' ,%s不可为空'
    temp_code = respcode
    temp = form_var
    # temp = eval(form_var)
    for k in param:
        if not temp.get(k):
            respcode, respmsg = str(temp_code), param.get(k) + '不可为空'
            # respinfo = HttpResponse(public.setrespinfo())
            return respcode, respmsg


form_var = {
    "supply_unit": "联桥科技有限公司",
    "rec_unit": "",
    "supply_address": "许昌市中原电气谷森尼瑞节能产业园四楼",
    "rec_address": "",
    "sp_contact": "",
    "rec_contact": "",
    "supply_phone": "",
    "rec_phone": "",
    "ERP_no": "",
    "express_no": "",
    "notice_date": "2021-04-16 14:36:55",
    "send_date": "",
    "order_detail": [],
    "express_req_options": [
        {
            "key": "HT",
            "value": "航天送货单"
        },
        {
            "key": "ZH",
            "value": "中慧平台送货单"
        },
        {
            "key": "LQ",
            "value": "联桥送货单"
        },
        {
            "key": "ZY",
            "value": "专用送货单"
        },
        {
            "key": "QT",
            "value": "其他"
        }
    ],
    "express_req": "HT",
    "special_req": None,
    "tab_man": "",
    "qa_man": "",
    "store_man": ""
}
param_not_null = {
    'supply_unit': '发货单位',
    'rec_unit': '收货单位',
    'supply_address': '发货方地址',
    'rec_address': '收货方地址',
    'sp_contact': '发货方联系人',
    'rec_contact': '收货方联系人',
    'supply_phone': '发货方电话',
    'rec_phone': '收货方电话',
    'ERP_no': 'ERP销货单号',
    'express_no': '货运单号',
    'notice_date': '通知时间',
    'send_date': '发货日期',
    'order_detail': '订单明细'
}


# not_null_response(1001, **param_not_null)
# print(readxlsx())
# a = public_db.Get_SeqNo("MEETING_ROOM")
# print(a)
# print(datetime.datetime.strptime('2022-01-07', '%Y-%m-%d').date().isocalendar()[1])
# dxweek(datetime.datetime.strptime('2021-03-15', '%Y-%m-%d').date())
# ws = inwb.get_sheet_by_name(sheetname1[0])
# rows = ws.max_row
# columns = ws.max_column
#
# print(ws.cell(1, 4).value)

# sql = "insert into yw_workflow_shipnotice (supply_unit, rec_unit, supply_address, rec_address, sp_contact, " \
#               "rec_contact, supply_phone, rec_phone, ERP_no, express_no, notice_date, send_date, order_detail," \
#               " express_req, special_req, tab_man, tran_date, check_state) " \
#               "values (%s, %s, %s,%s, %s, %s, %s, %s,%s, %s, %s, %s, %s,%s, %s, %s, %s,%s)"


def create_insert_value(sql):
    """根据insert语句创建对应value"""
    columns = sql[sql.index('(') + 1:sql.index(')')].replace(' ', '')
    columns_list = columns.split(',')
    sql_value = []
    for column in columns_list:
        temp_str = "form_var.get('%s')" % column
        sql_value.append(temp_str)
    sql_value = str(tuple(sql_value)).replace('"', '')
    return sql_value


sql = "insert into yw_workflow_chk_price (tab_no, tran_date, customer, prod_name, prod_type, prod_unit, " \
      "prod_count, remark_bom, amount_bom, remark_scfl, amount_scfl, remark_rgcb, amount_rgcb, " \
      "remark_utility, amount_utility, remark_depre, amount_depre, remark_loss, amount_loss, remark_manage, " \
      "amount_manage, remark_trans, amount_trans, remark_profit, amount_profit, remark_tax, amount_tax, " \
      "remark_market_price, market_price, remark_suggest_price, suggest_price, remark_confirm_price, " \
      "confirm_price, tab_man,check_state) values ()"

# print(create_insert_value(sql))

var_str = {
    "tab_no": "",
    "tran_date": "2021-04-17 13:08:08",
    "prod_name": "",
    "prod_type": "",
    "prod_unit": "",
    "prod_count": "",
    "remark_bom": "",
    "amount_bom": "",
    "remark_scfl": "/",
    "amount_scfl": "",
    "remark_rgcb": "工资",
    "amount_rgcb": "",
    "remark_utility": "0.6%*(材料成本+人工成本)",
    "amount_utility": "",
    "remark_depre": "0.9%*(材料成本+人工成本)",
    "amount_depre": "",
    "remark_loss": "2%*(材料成本+人工成本)",
    "amount_loss": "",
    "remark_manage": "1%*(材料成本+人工成本)",
    "amount_manage": "",
    "remark_trans": "/",
    "amount_trans": "",
    "remark_profit": "(BOM材料成本+生产辅材+人工成本+水电+折旧+损耗+管理+运输)*10%",
    "amount_profit": "",
    "remark_tax": "(人工成本+制造成本+管理费+运输+利润)*13%",
    "amount_tax": "",
    "remark_market_price": "",
    "market_price": "",
    "remark_suggest_price": "BOM材料成本+生产辅材+人工成本+水电+折旧+损耗+管理+运输+税费+利润",
    "suggest_price": "",
    "remark_confirm_price": "/",
    "confirm_price": "",
    "tab_man": "",
    "confirm_man": "",
    "check_man": ""
}
# list_param = ["`id` int(11) NOT NULL"]
# c_sql = "create table yw_workflow_chk_price "
#
# for param in var_str:
#     list_param.append(param + " varchar(255) NULL")
# c_sql += str(tuple(list_param))
# c_sql = c_sql.replace("'", '')
# print(c_sql)


digts = {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}


def char2num(c):
    return digts.get(c)


def str2num(s):
    """字符串转数字"""
    return reduce(lambda x, y: x * 10 + y, map(char2num, s))


def is_num(s):
    digts['.'] = '.'
    if isinstance(s, str):
        for i in s:
            if i in digts:
                pass
            else:
                return False
        return True
    elif isinstance(s, int) or isinstance(s, float):
        return True
    else:
        return False


def test1():
    l1 = [('a', 1), ('a', 2), ('a', 3), ('b', 10), ('b', 11), ('a', 4), ('d', 0xFF)]
    ln = [x for x in l1 if x[0] == 'a']
    lm = filter(lambda x: x[0] == 'a', l1)
    # print(list(lm))
    l1.sort(key=lambda x: x[0])
    c = itertools.groupby(l1, lambda x: x[0])
    lc = {}
    for k, v in c:
        print(list(v)[0])
        # # sum_value = 0
        # # for i in v:
        # #     sum_value += i[1]
        # sum_value = sum(x[1] for x in v)
        # lc[k] = sum_value
    print(lc)
    s = {}
    for k, v in itertools.groupby(l1, lambda x: x[0]):
        s[k] = sum(x[1] for x in v)
    print(s)


class nation():
    def __init__(self, key1, value1):
        self.key = key1
        self.value = value1

    def desc(self):
        return r"'key': '" + self.key + r"', 'value': '" + self.value + "'"


def test2():
    l1 = list(range(1, 10))
    tot = reduce(lambda x, y: x * y, l1)
    print(tot)
    l2 = map(lambda m: m ** 2, l1)
    print(list(l2))
    han = nation('01', '汉族')
    print(han.__dict__)
    print(han.desc())


def test3():
    file_path = r"C:\Users\user\Documents\a.json"
    lx = []
    with(open(file_path, encoding='utf-8')) as f:
        for line in f:
            # temp_row = 'key:'+line[:line.index(',')]
            mz = nation(line[:line.index(',')], line[line.index(',') + 1:-1])
            lx.append(mz.__dict__)
    print(lx)


def auto_generate_sql(input):
    temp_input = input[input.index('(') + 1:input.index(')')].replace(' ', '')
    list_param = temp_input.split(',')
    temp_input = ''
    ln = []
    if list_param:
        for item in list_param:
            form_value = r"form_var.get('%s')" % item

            ln.append(str(form_value))
    return str(tuple(ln)).replace('"', '')


sql = "insert into sys_staff_info (job_number, staff_name, name_spell, id_number, personal_phone, " \
      "company_phone, attend_number, gender, nation, birthday, address, hometown, degree, graduate_school, " \
      "graduate_date, major, politic_face, height, weight, health_state, marital_status, exp_salary, postbox, " \
      "emergency_contact, relationship, emergency_contact_phone, entry_date, work_state, less_id)" \
      "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
# print(auto_generate_sql(sql))

rows = [(1, 2), (3, 4), (5, 6)]
columns_list = map(lambda x: x[0], rows)
print(list(columns_list))

def insert_or_update_table(cur, tablename, **param):
    """根据表名更新或插入数据"""
    if not tablename:
        return None
    try:
        # cur = connection.cursor()
        # 根据表名查询字段和注释
        sql = "SELECT COLUMN_NAME,column_comment FROM INFORMATION_SCHEMA.Columns WHERE table_name= %s AND " \
              "table_schema= %s "
        sql_value = (tablename, 'lqkj_db')
        cur.execute(sql, sql_value)
        rows = cur.fetchall()
        if not rows:
            return None
        columns = map(lambda x: x[0], rows)
        if isinstance(param, dict):
            rp = set(param.keys()) & set(columns)
            # 根据id判断更新还是插入
            if 'id' in rp:
                sql = "select 1 from %s where id = %s"
                cur.execute(sql, (tablename, param.get('id')))
                row = cur.fetchone()
                if row:
                    # update
                    sql = "update " + tablename + " set "
                    for k in rp:
                        sql += k + '='
                        sql += param.get(k)
                    sql += "where id=" + param.get('id')
                else:
                    # insert
                    sql = "insert into " + tablename
                    sql += ' ' + str(list(rp))
                    sql += ' values ' + str(tuple(map(lambda x: param.get(x), list(rp))))
            else:
                # insert
                sql = "insert into " + tablename
                sql += ' ' + str(list(rp))
                sql += ' values ' + str(tuple(map(lambda x: param.get(x), list(rp))))
        # excute
        # return cur.execute(sql)
        cur.close()
        return sql
    except Exception as ex:
        return ex.__str__()


dc = {'row1': 'v1', 'row5': 'v3', 'row3': 'v5'}
lc = ['row2', 'row3', 'row5']

p = set(dc.keys()) & set(lc)
if 'row1' in p:
    print(True)
else:
    print(False)

test_var = {
  "absence_apply_table": None,
  "user_name": "吴晓萌",
  "org_options": [
    {
      "key": "XT",
      "value": "XT-系统部"
    }
  ],
  "department": "XT",
  "position_name": "default",
  "position_no": "",
  "position_number": 999,
  "onduty_number": 0,
  "min_salary": "0",
  "celling_wage": "99999",
  "position_duty": "默认岗位，请尽快调整到对应岗位",
  "tran_date": "2021-04-26 09:12:31",
  "order_number": None,
  "status_options": [
    {
      "key": "0",
      "value": "0-申请中"
    },
    {
      "key": "1",
      "value": "1-申请成功"
    },
    {
      "key": "2",
      "value": "2-申请失败"
    },
    {
      "key": "3",
      "value": "3-已归还"
    }
  ],
  "apply_status": "0",
  "apply_state_options": [
    {
      "key": "0",
      "value": "0-发起申请"
    },
    {
      "key": "1",
      "value": "1-审核通过"
    },
    {
      "key": "2",
      "value": "2-审核未通过"
    },
    {
      "key": "3",
      "value": "3-审批完结"
    },
    {
      "key": "5",
      "value": "5-等待审批"
    }
  ],
  "approve_status": "0",
  "id": None,
  "category_options": [
    {
      "key": 1,
      "value": "1-事假"
    },
    {
      "key": 2,
      "value": "2-调休"
    },
    {
      "key": 3,
      "value": "3-产假"
    },
    {
      "key": 4,
      "value": "4-陪产假"
    },
    {
      "key": 5,
      "value": "5-婚假"
    },
    {
      "key": 6,
      "value": "6-丧假"
    }
  ],
  "category": "",
  "start_date": "",
  "days": "",
  "agent_name": "",
  "end_date": "",
  "reason": "",
  "remark": "",
  "status": "0",
  "apply_state": "0",
  "user_id": ""
}


# temp_var = {k: (v if v else '') for k, v in test_var.items()}
# for k, v in temp_var.items():
#     print(k, v)

def rowlist2dict(ln):
    tempset = {}
    for k, *v in ln:
        tempset[k] = v
    return tempset


def test_row2dict():
    ln = [('user1', 'name1', 'age1'), ('user2', 'name2', 'age2'), ('user3', None, 'age3')]
    for k, v in rowlist2dict(ln).items():
        print(k, v)


ln = [('user1', 'name1'), ('user2', 'name2'), ('user3', None)]
t = {k: v for k, v in ln}
tit = ['user', 'name']
lt = [{'user': k, 'name': v} for k, v in ln]
col = {'user1': 'varchar', 'age1': 'int', 'user2': 'float', 'user3': 'double'}
print(col.__doc__)
dict_list = []
for cn, cs in ln:
    tempdct = {'class_name': cn, 'class_spell': cs}
    dict_list.append(tempdct)
print(dict_list)
seclit = [
  {
    "class_name": "11",
    "class_spell": "122"
  },
  {
    "class_name": "22",
    "class_spell": "33"
  }
]
for item in seclit:
    cn = item.get('class_name')
    cs = item.get('class_spell')
    print(cn, cs)

s = 'BD_SE_TD'
if 'd' in s:
    print(True)
s += '_' * 2
ts = s.split('_')
company, seclass, trdclss, *_ = ts
print(company, seclass, trdclss, _)

column_dct = [{k: v} for k, v in rows]
print([x.keys() for x in column_dct])

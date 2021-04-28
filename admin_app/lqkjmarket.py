from django.shortcuts import HttpResponse
from django.db import connection
import json
from admin_app import public
from admin_app import models
import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment,Border,Side

number_keys=[
            'sale_money_year',  # 年销货金额
            'not_invoice_money_year',  # 年未开票金额
            'invoice_money_year',  # 年开票金额
            'year_return_money',  # 年回款金额
            'need_money',  # 应收款
            'need_money_endtime',  # 到期应收款
            'need_money_not_endtime',  # 未到期应收款
            'return_money_month',  # 本月回款金额
            'not_found_money',  # 未识别金额
            'order_all_amount',#总订单金额
            'year_not_return_money'#总未回款金额
        ]


def main(request):
    log = public.logger
    log.info('----------------------lqkjmarket-begin---------------------------')
    if request.method == "POST": #POST请求
        #请求body转为json
        tmp=request.body
        tmp=tmp.decode(encoding='utf-8')
        request_body=json.loads(tmp)
        trantype=request_body['trantype']
        log.info('trantype=[%s]' % trantype)
        fun_name = trantype
        if globals().get(fun_name):
            resp = globals()[fun_name](request, request_body)
        else:
            s = public.setrespinfo({"respcode": "100000", "respmsg": "api error! lqkjerp!"})
            resp = HttpResponse(s)
    elif request.method == "GET": #GET请求
        s = public.setrespinfo({"respcode": "100000", "respmsg": "api error"})
        resp = HttpResponse(s)
    log.info('----------------------lqkjmarket-end---------------------------')
    return resp

#函数装饰器
def func_wrapper(func):
    def wrapper(request, request_body,*args,**kwargs):
        log = public.logger
        fun_name=func.__name__
        log.info('----------------------lqkjmarket-%s-begin---------------------------'%fun_name)
        jsondata = {
            "respcode": "000000",
            "respmsg": "交易成功",
            "trantype": request_body.get('trantype', None),
        }
        res=func(jsondata,log,request, request_body,*args,**kwargs)
        s = json.dumps(jsondata, cls=models.JsonCustomEncoder, ensure_ascii=False)
        log.info(s)
        log.info('----------------------lqkjmarket-%s-end---------------------------'%fun_name)
        return HttpResponse(s)
    return wrapper


############################################3


#######################################################################

#查询未到期质保金
def get_qua_money_not_end_time(cursor,log,cre_id,end_date):
    ##查询未到期质保金
    sql = "SELECT IFNULL(sum(A.order_qua_deposit),0) FROM	yw_rece2_order_info_sub A LEFT JOIN yw_rece2_order_info B ON A.order_id = B.order_id where B.order_cre_id = %s  and B.order_paytype='1' AND date_add(	A.send_date,INTERVAL A.order_bao_term DAY) > %s"
    log.info(sql % (cre_id, end_date))
    cursor.execute(sql, (cre_id, end_date))
    row = cursor.fetchone()
    if row:
        qua_money = row[0]
    else:
        qua_money = 0
    return qua_money
#获取到期应收款
def get_need_money_end_time(cursor,log,cre_id,begin_date,end_date,return_money):
    ##查询货到付款应收款金额
    sql = "SELECT IFNULL(sum(A.order_amount),0) FROM yw_rece2_order_info_sub A LEFT JOIN yw_rece2_order_info B ON A.order_id = B.order_id where B.order_cre_id = %s and B.order_paytype='1' AND date_add(	A.send_date,INTERVAL B.order_assperiod DAY) >= %s and date_add(	A.send_date,INTERVAL B.order_assperiod DAY)<=%s"
    log.info(sql % (cre_id, begin_date,end_date))
    cursor.execute(sql, (cre_id, begin_date,end_date))
    row = cursor.fetchone()
    if row:
        need_money_endtime_1 = row[0]
    else:
        need_money_endtime_1 = 0
    ##查询票到付款应收款金额
    sql = "SELECT IFNULL(sum(A.code_amo_money),0) FROM yw_rece2_invoice_info A LEFT JOIN yw_rece2_order_info B ON SUBSTRING_INDEX(A.con_order_id,',',1) = B.order_id where B.order_cre_id = %s and B.order_paytype='2' AND date_add(	A.code_date,INTERVAL B.order_assperiod DAY) >= %s and date_add(	A.code_date,INTERVAL B.order_assperiod DAY)<=%s"
    log.info(sql % (cre_id, begin_date,end_date))
    cursor.execute(sql, (cre_id, begin_date,end_date))
    row = cursor.fetchone()
    if row:
        need_money_endtime_2 = row[0]
    else:
        need_money_endtime_2 = 0
    ##查询扣除总金额
    sql="select IFNULL(sum(order_ded_money),0) from yw_rece2_order_info where order_cre_id='%s'"%cre_id
    log.info(sql)
    cursor.execute(sql)
    row = cursor.fetchone()
    if row:
        ded_money = row[0]
    else:
        ded_money = 0
    need_money_endtime = float(need_money_endtime_1) + float(need_money_endtime_2)-ded_money

    ##查询未到期质保金
    qua_money=get_qua_money_not_end_time(cursor,log,cre_id,end_date)
    # sql = "SELECT IFNULL(sum(A.order_qua_deposit),0) FROM	yw_rece2_order_info_sub A LEFT JOIN yw_rece2_order_info B ON A.order_id = B.order_id where B.order_cre_id = %s  and B.order_paytype='1' AND date_add(	A.send_date,INTERVAL A.order_bao_term DAY) > %s"
    # log.info(sql % (cre_id, format_date))
    # cursor.execute(sql, (cre_id, format_date))
    # row = cursor.fetchone()
    # if row:
    #     qua_money = row[0]
    # else:
    #     qua_money = 0

    #到期应收款=到期应收款-未到期质保金-已回款
    need_money_endtime = float(need_money_endtime) - float(qua_money) - float(return_money)
    return need_money_endtime

#获取未到期应收款
def get_need_money_not_end_time(cursor,log,cre_id,end_date):
    ##查询货到付款未到期收款金额
    sql = "SELECT IFNULL(sum(A.order_amount),0) FROM yw_rece2_order_info_sub A LEFT JOIN yw_rece2_order_info B ON A.order_id = B.order_id where B.order_cre_id = %s and B.order_paytype='1' AND date_add(	A.send_date,INTERVAL B.order_assperiod DAY) > %s"
    log.info(sql % (cre_id, end_date))
    cursor.execute(sql, (cre_id, end_date))
    row = cursor.fetchone()
    if row:
        need_money_not_endtime_1 = row[0]
    else:
        need_money_not_endtime_1 = 0
    ##查询票到付款未到期收款金额
    sql = "SELECT IFNULL(sum(A.code_amo_money),0) FROM yw_rece2_invoice_info A LEFT JOIN yw_rece2_order_info B ON SUBSTRING_INDEX(A.con_order_id,',',1) = B.order_id where B.order_cre_id = %s and B.order_paytype='2' AND date_add(	A.code_date,INTERVAL B.order_assperiod DAY) > %s"
    log.info(sql % (cre_id, end_date))
    cursor.execute(sql, (cre_id, end_date))
    row = cursor.fetchone()
    if row:
        need_money_not_endtime_2 = row[0]
    else:
        need_money_not_endtime_2 = 0
    need_money_not_end_time=float(need_money_not_endtime_1)+float(need_money_not_endtime_2)
    return need_money_not_end_time

def _formatDate(date):
    fmt = '%Y-%m-%d %H:%M:%S'
    time_tuple = time.strptime(date, fmt)
    format_date = datetime.datetime(*time_tuple[:6])
    return format_date


@func_wrapper
def get_excel_data(jsondata,log,request, request_body):
    date=request_body.get('dateinterval')
    if not date:
        jsondata['respcode'] = '400000'
        jsondata['respmsg'] = '缺少必要参数'
        return
    begin_date,end_date=date
    begin_cus=request_body.get('begin_cus')
    end_cus=request_body.get('end_cus')

    if not begin_date or not end_date:
        jsondata['respcode'] = '400000'
        jsondata['respmsg'] = '缺少必要参数'
        return
    fmt = '%Y-%m-%d %H:%M:%S'
    begin_date=_formatDate(begin_date)
    end_date=_formatDate(end_date)
    # next_month_end_day = public.get_next_month_end_day(format_date)#下月最后一天

    cursor = connection.cursor()
    #查询指定客户信息
    if begin_cus and end_cus:
        sql="select cus_code,cus_rank,cus_name,cre_id from yw_rece2_customer_info where cus_code>=%s and cus_code<=%s"
        log.info(sql%(begin_cus,end_cus))
        cursor.execute(sql,(begin_cus,end_cus))
    else:#查询所有客户信息
        sql = "select cus_code,cus_rank,cus_name,cre_id from yw_rece2_customer_info"
        log.info(sql)
        cursor.execute(sql)
    curs = cursor.fetchall()
    detail=[]
    for cus_code,cus_rank,cus_name,cre_id in curs:
        cus={}
        cus['cus_code']=cus_code
        cus['cus_rank']=cus_rank
        cus['cus_name']=cus_name
        cus['cre_id']=cre_id

        #查询总开票金额
        sql="select IFNULL(sum(code_amo_money),0) from yw_rece2_invoice_info where cre_id=%s and code_date>=%s and code_date<=%s"
        log.info("查询总开票金额:"+sql%(cre_id,begin_date,end_date))
        cursor.execute(sql,(cre_id,begin_date,end_date))
        row=cursor.fetchone()
        if row and row[0]:
            cus['year_invoice_money']=row[0]
        else:
            cus['year_invoice_money'] = 0
        invoice_money = cus['year_invoice_money']  #总开票金额
        invoice_money_year = cus['year_invoice_money']  #总开票金额

        #查询总回款金额
        sql="select IFNULL(sum(ret_amo_money),0) from yw_rece2_return_info where cre_id=%s and ret_date>=%s and ret_date<=%s"
        log.info("总回款金额:"+sql%(cre_id,begin_date,end_date))
        cursor.execute(sql,(cre_id,begin_date,end_date))
        row = cursor.fetchone()
        if row and row[0]:
            cus['year_return_money'] = row[0]
        else:
            cus['year_return_money'] = 0
        # 总回款金额
        return_money = cus['year_return_money']

        #查询发货金额
        #sql = "select IFNULL(sum(A.order_amount),0) from yw_rece2_order_info_sub as A left JOIN yw_rece2_order_info as B on A.order_id=B.order_id where B.order_cre_id=%s and B.order_paytype='1' and A.send_date>=%s and A.send_date<=%s"
        sql = "select IFNULL(sum(order_amount),0) from yw_rece2_order_info_sub where cre_id='%s' and send_date>='%s' and send_date<='%s'" % (cre_id,begin_date,end_date)
        log.info("查询发货金额:"+sql)
        cursor.execute(sql)
        row = cursor.fetchone()
        if row and row[0]:
            send_money = row[0]
        else:
            send_money = 0
        cus['sale_money_year'] = send_money #总销货金额

        #应收款金额=开票金额-回款金额
        #cus['need_money'] = float(invoice_money)+float(send_money)-float(return_money)
        cus['need_money'] = float(invoice_money) - float(return_money)

        #查询到期应收款
        ##查询货到付款应收款金额
        # sql="SELECT IFNULL(sum(order_amount),0) FROM yw_rece2_order_info_sub where cre_id = %s AND date_add( send_date,INTERVAL B.order_assperiod DAY) < %s"
        # log.info("查询货到付款应收款金额:"+sql%(cre_id,end_date))
        # cursor.execute(sql,(cre_id,end_date))
        # row = cursor.fetchone()
        # if row and row[0]:
        #     need_money_endtime_1 = row[0]
        # else:
        #     need_money_endtime_1 = 0
        ##查询票到付款应收款金额
        sql = "select IFNULL(sum(code_amo_money),0) from yw_rece2_invoice_info where cre_id='%s' " \
              "AND date_add( code_date,INTERVAL order_assperiod DAY) < '%s'" % (cre_id, end_date)
        log.info("查询票到付款应收款金额:"+sql)
        cursor.execute(sql)
        row = cursor.fetchone()
        if row and row[0]:
            need_money_endtime_2 = row[0]
        else:
            need_money_endtime_2 = 0

        #查询未到期质保金
        sql = "select IFNULL(sum(order_qua_deposit),0) from yw_rece2_order_info_sub where cre_id='%s' " \
              "AND date_add( send_date,INTERVAL order_bao_term DAY) > '%s'" % (cre_id, end_date)
        log.info("查询未到期质保金:" + sql)
        cursor.execute(sql)
        row = cursor.fetchone()
        if row and row[0]:
            qua_money = row[0]
        else:
            qua_money = 0

        #到期应收款=到期应收款-未到期质保金-已回款
        cus['need_money_endtime'] = float(need_money_endtime_2) - float(return_money) - float(qua_money) #到期应收款=到期总应收款-已回款-未到期质保金

        #查询未到期应收款
        cus['need_money_not_endtime'] = float(invoice_money) +float(qua_money) - float(need_money_endtime_2) #总货/票款+ 未到期质保金- 到期应收款

        #查询本月回款金额
        sql="select IFNULL(sum(ret_amo_money),0) from yw_rece2_return_info where cre_id=%s and YEAR(ret_date)=YEAR(%s) and MONTH(ret_date)=MONTH(%s)"
        log.info("查询本月回款金额:" +  sql%(cre_id,end_date,end_date))
        cursor.execute(sql,(cre_id,end_date,end_date))
        row = cursor.fetchone()
        if row and row[0]:
            cus['return_money_month'] = row[0]
        else:
            cus['return_money_month'] = 0

        #查询未识别金额（回款表里订单为空的记录金额和）
        sql="select IFNULL(sum(ret_amo_money),0) from yw_rece2_return_info where cre_id=%s and (con_order_id is null or con_order_id='') and ret_date>=%s and ret_date<=%s"
        log.info(sql,(cre_id,begin_date,end_date))
        cursor.execute(sql,(cre_id,begin_date,end_date))
        row = cursor.fetchone()
        if row:
            cus['not_found_money'] = row[0]
        else:
            cus['not_found_money'] = 0

        #查询总未开票金额
        cus['invoice_money_year']=invoice_money_year
        ##总未开票金额=总销货金额-总开票金额
        cus['not_invoice_money_year']=float(cus['sale_money_year']) - float(invoice_money_year)
        cus['invoice_money_next_month_plan']=''


        #20200111新增
        # 查询总订单金额
        sql = "select IFNULL(sum(order_amount),0) from yw_rece2_order_info where order_cre_id=%s and order_tran_date>=%s and order_tran_date<%s"
        # log.info(sql, (cre_id, begin_date, end_date))
        cursor.execute(sql, (cre_id, begin_date, end_date))
        row = cursor.fetchone()
        if row:
            cus['order_all_amount'] = row[0]
        else:
            cus['order_all_amount'] = 0

        # 查询总未回款金额
        ## 总未回款=总销货-总回款
        cus['year_not_return_money']=cus['sale_money_year']-cus['year_return_money']

        #备注
        cus['remark']=''

        #判断如果数值全0不统计
        flag=False
        for key,value in cus.items():
            if key in number_keys:
                #如果有不等于0的 flag=True
                if float(value)!=float(0):
                    flag=True
                    #保留6位小数
                    cus[key]=round(value,6)
        #如果全0 跳过
        log.info('flag=%s'%flag)
        if not flag:
            continue

        detail.append(cus)

    cursor.close() #关闭游标
    #返回报文结构体
    jsondata['tableData']=detail
    jsondata['tableHead']=[
        {
            'name': 'cus_code',
            'label': '客户代号',
            'width': 150
        },
        {
            'name': 'cus_rank',
            'label': '客户等级',
            'width': 150
        },
        {
            'name': 'cus_name',
            'label': '客户名称',
            'width': 150
        },
        {
            'name': 'order_all_amount',
            'label': '总订单金额',
            'width': 150
        },
        {
            'name': 'sale_money_year',
            'label': '总销货金额',
            'width': 150
        },
        {
            'name': 'not_invoice_money_year',
            'label': '总未开票金额',
            'width': 150
        },
        {
            'name': 'invoice_money_year',
            'label': '总开票金额',
            'width': 150
        },
        {
            'name': 'year_return_money',
            'label': '总回款金额',
            'width': 150
        },
        {
            'name': 'year_not_return_money',
            'label': '总未回款金额',
            'width': 150
        },
        {
            'name': 'need_money',
            'label': '应收款',
            'width': 150
        },
        {
            'name': 'need_money_endtime',
            'label': '到期应收款',
            'width': 150
        },
        {
            'name': 'need_money_not_endtime',
            'label': '未到期应收款',
            'width': 150
        },
        {
            'name': 'return_money_month',
            'label': '本月回款金额',
            'width': 150
        },
        {
            'name': 'invoice_money_next_month_plan',
            'label': '下月预计回款金额',
            'width': 150
        },
        {
            'name': 'not_found_money',
            'label': '未识别金额',
            'width': 150
        },
        {
            'name': 'remark',
            'label': '备注',
            'width': 150
        },


    ]
    if begin_cus and end_cus:
        filename = '联桥科技应收账款_%s-%s_%s-%s.xlsx' % (begin_cus,end_cus,begin_date.strftime('%Y%m%d%H%M%S'), end_date.strftime('%Y%m%d%H%M%S'))
    else:
        filename='联桥科技应收账款_%s-%s.xlsx'%(begin_date.strftime('%Y%m%d%H%M%S'),end_date.strftime('%Y%m%d%H%M%S'))
    _make_excel(jsondata['tableHead'],jsondata['tableData'],filename)
    jsondata['fileurl']='http://10.10.10.174/static/filedown/%s'%filename


def _getGbkLen(w):
    return len(str(w).encode('gbk'))

def _make_excel(tableHead,tableData,filename):
    # 拼装数据
    fullData = []
    for i in range(len(tableHead)):
        temp = []
        label = tableHead[i]['label']
        name = tableHead[i]['name']
        temp.append(label)
        for j in range(len(tableData)):
            temp.append(tableData[j][name])
        if name == 'cus_code':
            temp.append('总计：')
        elif name in number_keys:
            temp.append(sum(temp[1:]))
        else:
            temp.append('')
        fullData.append(temp)

    wb = Workbook()
    mySheet = wb.create_sheet(index=0, title="Mysheet")

    fontStyle = Font(size=18)
    fontStyle2 = Font(size=14)
    fontMini = Font(size=10)
    centerStyle = Alignment(horizontal='center', vertical='center')
    borderStyle = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')

    )

    startLetter = chr(65)
    endLetter = chr(65 + len(tableHead) - 1)
    mySheet.merge_cells('%s1:%s1' % (startLetter, endLetter))
    mySheet.merge_cells('%s2:%s2' % (startLetter, endLetter))
    mySheet.cell(1, 1).value = '联桥科技有限公司'
    mySheet.cell(1, 1).alignment = centerStyle
    mySheet.cell(1, 1).font = fontStyle
    mySheet.cell(2, 1).value = '应收账款'
    mySheet.cell(2, 1).alignment = centerStyle
    mySheet.cell(2, 1).font = fontStyle2

    mySheet.cell(3, len(tableHead) - 2).value = '打印日期：'
    mySheet.cell(3, len(tableHead) - 2).alignment = centerStyle
    mySheet.cell(3, len(tableHead) - 2).font = fontMini

    row = 0
    column = 1
    for items in fullData:
        row = 4
        tempSorted = sorted(items, key=lambda x: -_getGbkLen(x))
        maxlen = _getGbkLen(tempSorted[0]) + 4

        # 调整列高
        mySheet.column_dimensions[chr(65 + column - 1)].width = maxlen
        for item in items:
            # 调整行高
            mySheet.row_dimensions[row].height = 25
            mySheet.cell(row, column).value = item
            mySheet.cell(row, column).alignment = centerStyle
            mySheet.cell(row, column).border = borderStyle
            row += 1
        column += 1

    row += 1
    column = 2
    mySheet.cell(row, column).value = '审核：'
    mySheet.cell(row, column).alignment = centerStyle

    column = int(len(tableHead) / 2)
    mySheet.cell(row, column).value = '制表人：'
    mySheet.cell(row, column).alignment = centerStyle

    column = len(tableHead) - 2
    mySheet.cell(row, column).value = '时间：'
    mySheet.cell(row, column).alignment = centerStyle


    wb.save(public.localhome+'static/filedown/%s'%filename)
    wb.close()
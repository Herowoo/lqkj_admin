# import xlrd
# from xlutils.copy import copy
import openpyxl
import shutil
from openpyxl.styles import Border,Side
from copy import copy

class TempRender:
    def __init__(self,temp_path,save_path):
        """
        :param temp_path: 模板路径
        """
        self.save_path=save_path
        shutil.copy(temp_path,save_path)
        self.wb=openpyxl.load_workbook(temp_path)
        self.ws=self.wb[self.wb.sheetnames[0]]

        self.wb2=openpyxl.load_workbook(save_path)
        self.ws2=self.wb2[self.wb2.sheetnames[0]]



    def render(self,kvs):
        """渲染"""
        render_data=[]
        end_row=0
        # new_col=0
        print(self.ws.max_row,self.ws.max_column)
        add_row=[0 for i in range(self.ws.max_row+1)]
        add_col = [0 for i in range(self.ws.max_column+1)]
        for i in range(self.ws.max_column):
            temp=[]
            for j in range(self.ws.max_row):
                value=self.ws.cell(j+1,i+1).value
                if type(value)==str and ('$' in value or '#' in value):
                    # 渲染纵向列表
                    if '$[' in value and ']' in value:
                        key=value.replace('$[','').replace(']','')
                        if key in kvs:
                            #写渲染table列
                            self.render_write_down(j+add_col[i+1],i+add_row[j+1],kvs[key])
                            add_col[i+1]+=len(kvs[key])-1
                            continue
                    # 渲染横向列表
                    if '#[' in value and ']' in value:
                        key = value.replace('#[', '').replace(']', '')
                        if key in kvs:
                            # 写渲染table行
                            self.render_write_right(j + add_col[i+1], i + add_row[j+1], kvs[key])
                            add_row[i+1]+=len(kvs[key]) - 1

                            continue
                    #渲染sum函数
                    if '$sum(' in value and ')' in value:
                        key=value.replace('$sum(','').replace(')','')
                        if key in kvs:
                            self.ws2.cell(j+1 + add_col[i+1], i+1+add_row[j+1], sum(kvs[key]))
                            continue
                #不渲染 直接填充
                if value:
                    print(i,j,value,add_row,add_col)
                    # print(add_col,add_row)
                    self.ws2.cell(j+1 + add_col[i+1], i+1 + add_row[j+1], value)
                temp.append(value)
            render_data.append(temp)
        print(render_data)

    def render_write_right(self,pos_row,pos_col,data):
        """
        向右渲染写入excel
        :param pos_row: 
        :param pos_col: 
        :param data: 
        :return: 
        """
        print('渲染行',locals())
        style = self.ws.cell(pos_row + 1, pos_col + 1)._style
        for item in data:
            self.ws2.cell(pos_row+1,pos_col+1,item)
            self.ws2.cell(pos_row + 1, pos_col + 1)._style = copy(style)
            pos_col+=1

    def render_write_down(self,pos_row,pos_col,data):
        """
        向下纵向渲染写入excel
        :param pos_row: 起始行位置
        :param pos_col: 起始列位置
        :param data: 
        :return: 
        """
        style=self.ws.cell(pos_row+1,pos_col+1)._style
        for item in data:
            # self.new_ws.write(pos_row,pos_col,item)
            self.ws2.cell(pos_row+1,pos_col+1).value=item
            self.ws2.cell(pos_row + 1, pos_col + 1)._style = copy(style)
            pos_row+=1


    def save(self):
        """保存文件"""
        self.wb2.save(self.save_path)


if __name__=='__main__':
    tr=TempRender('temp.xlsx','render_result.xlsx')
    kvs={
        'index':[1,2,3,4,5],
        'name':[ '陈志广', '魏改', '姚琼琼', '杨建涛', '姜晓旭'],
        'depart':['总经办', '财务部', '财务部', '企管部', '企管部'],
        'post':['总经理', '财务经理', '出纳', '主管', '人事专员'],
        'age':[1,2,3,4,5],
        'table_head':['序号','部门','职位','姓名','身份证号','年龄','民族','入职时间','学位',	'毕业学校','政治面貌','婚姻状况'	,'地址','个人电话','公司电话','邮箱']
    }
    tr.render(kvs)
    tr.save()

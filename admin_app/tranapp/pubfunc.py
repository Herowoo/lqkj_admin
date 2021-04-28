import datetime
import time
import calendar
# import pymysql
import openpyxl
from functools import reduce

from admin_app import models
# from admin_app.tranapp import pubfunc

dic_weekday = {1: '周一', 2: '周二', 3: '周三', 4: '周四', 5: '周五', 6: '周六', 7: '周日'}
# connection = pymysql.connect(host='10.10.10.174', user='lqkj', password='LQkj666_2019', db='lqkj_db')
connection = ''

def set_cross(sa, sb):
    """判断两个集合是否相交"""
    if max(sa) > min(sb) and min(sa) < max(sb):
        return True
    else:
        return False


def set_dif(a, b):
    """返回两个集合的差集"""
    if set_cross(a, b):
        if max(b) >= max(a) >= min(b) > min(a):
            t = min(b) - min(a)
        elif max(a) > max(b) >= min(a) >= min(b):
            t = max(a) - max(b)
        elif max(b) < max(a) and min(b) > min(a):
            t = max(a) - max(b)
            t += min(b) - min(a)
        else:
            t = min(b) - min(b)
    else:
        t = max(a) - min(a)
    return t


def set_inter(a, b):
    """返回两个集合交集"""
    if max(b) >= max(a) >= min(b) >= min(a):
        t = max(a) - min(b)
    elif max(a) >= max(b) >= min(a) >= min(b):
        t = max(b) - min(a)
    elif max(b) < max(a) and min(b) > min(a):
        t = max(b) - min(b)
    else:
        t = min(b) - min(b)
    return t


# 返回开始日期到截止日期的日期集合
def get_datelist(st, ed):
    dt_st = datetime.datetime.strptime(st, '%Y-%m-%d')
    dt_ed = datetime.datetime.strptime(ed, '%Y-%m-%d')
    dt = dt_st
    list_st = []
    while dt < dt_ed:
        dt = dt + datetime.timedelta(days=1)
        list_st.append(dt.strftime('%Y-%m-%d %H:%M:%S'))
    return list_st


# 返回日期相差天数
def get_days(st, ed):
    dt_st = time.strptime(st, '%Y-%m-%d %H:%M:%S')
    dt_ed = time.strptime(ed, '%Y-%m-%d %H:%M:%S')
    dt_st = datetime.datetime(dt_st[0], dt_st[1], dt_st[2])
    dt_ed = datetime.datetime(dt_ed[0], dt_ed[1], dt_ed[2])
    dt_dif = dt_ed - dt_st
    return dt_dif.days


def get_month_day_list(month_id):
    """根据月份返回当月日期列表,入参格式：yyyy-mm"""
    year = int(month_id[:4])
    month = int(month_id[5:])
    list_days = calendar.monthrange(year, month)
    last_day = list_days[1]
    begin_date = datetime.datetime.strptime(month_id, "%Y-%m")
    end_date = begin_date.replace(day=last_day)
    date_list = []
    while begin_date <= end_date:
        date_str = begin_date.strftime('%Y-%m-%d')
        weekday_str = dic_weekday.get(begin_date.isoweekday())
        date_tp = ['{:02d}'.format(begin_date.day), date_str, weekday_str]
        date_tp = tuple(date_tp)
        date_list.append(date_tp)
        begin_date += datetime.timedelta(days=1)
    return date_list


def getBetweenDay(begin_date):
    date_list = []
    begin_date = datetime.datetime.strptime(begin_date, "%Y-%m")
    end_date = begin_date.replace(day=calendar.monthrange())
    while begin_date <= end_date:
        date_str = begin_date.strftime("%Y-%m-%d")
        date_list.append(date_str)
        begin_date += datetime.timedelta(days=1)
    return date_list


def add_months(dt, months):
    month = dt.month - 1 + months
    year = int(dt.year + month / 12)
    month = month % 12 + 1
    day = dt.day
    last_day = calendar.monthrange(year, month)[1]
    day = min(day, last_day)
    return dt.replace(year=year, month=month, day=day)


def getBetweenMonth(begin_date, end_date):
    date_list = []
    begin_date = datetime.datetime.strptime(begin_date, "%Y-%m")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m")
    while begin_date <= end_date:
        date_str = begin_date.strftime("%Y-%m")
        date_list.append(date_str)
        begin_date = add_months(begin_date, 1)
    return date_list


def str_datetime(date_string, form_str='%Y-%m-%d %H:%M'):
    dt = datetime.datetime.strptime(date_string, form_str)
    return dt


# 返回所有加班时间（单位：秒）
def overtime_all(st, ed, uid):
    seconds = 0
    if get_days(st, ed) > 0:
        list_date = get_datelist(st[:10], ed[:10])
        i = 0
        while i < len(list_date) - 1:
            seconds += overtime_oneday(list_date[i], list_date[i][:10] + ' 23:59:59', uid)
            i += 1
        lastday = list_date[len(list_date) - 1]
        # 追加第一天加班时间
        seconds += overtime_oneday(st, st[:10] + ' 23:59:59', uid)
        # 追加最后一天加班时间
        seconds += overtime_oneday(lastday, ed, uid)
    else:
        seconds = overtime_oneday(st, ed, uid)
    return seconds


# 返回当天加班时间（单位：秒）
def overtime_oneday(st, ed, uid):
    fmt_str = "%Y-%m-%d %H:%M:%S"
    dt_st = datetime.datetime.strptime(st, fmt_str)
    dt_ed = datetime.datetime.strptime(ed, fmt_str)
    if is_holiday(st, uid) == 1:
        work = set([datetime.datetime.strptime("%s %s" % (st[:10], "08:30:00"), fmt_str),
                    datetime.datetime.strptime("%s %s" % (st[:10], "18:00:00"), fmt_str)])
    else:
        work = set([datetime.datetime.strptime("%s %s" % (st[:10], "11:30:00"), fmt_str),
                    datetime.datetime.strptime("%s %s" % (st[:10], "13:00:00"), fmt_str)])
    if dt_ed < dt_st:
        return 0
    else:
        return set_dif([dt_st, dt_ed], work).seconds


def is_holiday(st, uid):
    """是否节假日"""
    cur = connection.cursor()
    sql = "select group_id,user_id from yw_workflow_attend_group_user"
    cur.execute(sql)
    rows = cur.fetchall()
    group_list = []
    for row in rows:
        if uid in eval(row[1]):
            group_list.append(row[0])
    # 公休日列表
    res_day_list = []
    if group_list:
        sql = "select rest_day from yw_workflow_attend_group_cfg where group_id = %s and month_id = %s"
        for gl in group_list:
            cur.execute(sql, (gl, st[:7]))
            row = cur.fetchone()
            res_day_list += eval(row[0])
    if st[8:] in res_day_list:
        return True
    else:
        return False


def get_list_from_dict(parm_dict, dict_key):
    temp_list = []
    for item in parm_dict:
        temp_list.append(item.get(dict_key))
    return temp_list


def get_def_list(list_a, list_b):
    set_a = set(list_a)
    set_b = set(list_b)
    set_dif = set_a.difference(set_b)
    temp_list = list(set_dif)
    temp_list.sort()
    return temp_list


def get_insec_list(list_a, list_b):
    set_a = set(list_a)
    set_b = set(list_b)
    set_com = set_a.intersection(list_b)
    temp_list = list(set_com)
    temp_list.sort()
    return temp_list


def get_work_rest_day(user_id, month_id):
    """获取特定用户的工作日和休息日"""
    try:
        cur = connection.cursor()
        sql = "SELECT group_id, user_id FROM yw_workflow_attend_group_user"
        cur.execute(sql)
        rows = cur.fetchall()
        group_list = []
        for row in rows:
            if user_id == row[1]:
                group_list.append((row[0], month_id))
        if not group_list:
            return None
        sql = "SELECT work_day,rest_day from yw_workflow_attend_group_cfg where group_id = %s and month_id = %s"
        cur.executemany(sql, group_list)
        rows = cur.fetchall()
        list_work_day = []
        list_rest_day = []
        for row in rows:
            list_work_day.append(row[0])
            list_rest_day.append(row[1])
        tuple_work_day = set(list_work_day)
        tuple_rest_day = set(list_rest_day)
        tuple_work_day.sort()
        tuple_rest_day.sort()
        tuple_work_day = tuple(tuple_work_day)
        tuple_rest_day = tuple(tuple_rest_day)
        return tuple_work_day, tuple_rest_day
    except Exception as ex:
        pass
    finally:
        cur.execute("commit")
        cur.close()


def auto_generate_sql(input):
    temp_input = input.replace(r'=%s,', '')
    list_param = temp_input.split(',')
    temp_input = ''
    if list_param:
        for item in list_param:
            form_value = 'form_var.get(\'%s\'),' % item
            list_param += form_value
    return temp_input


def export_excel(filename, sql_select):
    wb = openpyxl.workbook.Workbook()
    sheet_1 = wb.active
    # sheet_1.title = "考勤报表"

    def get_columns_tablename_from_sql(sql_str):
        t_name = ''
        columns_list = []
        sql_str_lower = sql_str.lower()
        from_index = sql_str_lower.find('from')
        if from_index != -1:
            temp = sql_str_lower[from_index+4:]
            where_index = temp.find('where')
            # 不包含"where"子句
            if where_index == -1:
                pass
            else:
                temp = temp[:where_index]

            t_name = temp.replace(' ', '')
        # 截取列名字段
        select_index = sql_str_lower.find('select')
        if select_index != -1:
            temp = sql_str_lower[select_index+6:from_index]
            temp = temp.replace(' ', '')
            if temp != '*':
                columns_list = temp.split(',')
            else:
                # 获取所有列名
                pass
        return t_name, columns_list
    try:
        cur = connection.cursor()
        # 获取查询字段列表
        table_name, column_list = get_columns_tablename_from_sql(sql_select)
        sql_column_name_comment = "SELECT COLUMN_NAME,column_comment FROM INFORMATION_SCHEMA.Columns " \
                                  "WHERE table_name=%s"
        cur.execute(sql_column_name_comment, table_name)
        rows = cur.fetchall()
        dict_coloums = {}
        for row in rows:
            dict_coloums[row[0]] = row[1]
        title_list = []
        for column in column_list:
            title_list.append(dict_coloums.get(column))
        sheet_1.append(title_list)
        cur.execute(sql_select)
        rows = cur.fetchall()
        for row in rows:
            sheet_1.append(row)
        cur.close()
    except Exception as ex:
        print('Error:', ex)
    finally:
        wb.save(filename)
        wb.close()


def list2excel(filename, lst):
    wb = openpyxl.workbook.Workbook()
    sheet_1 = wb.active
    # sheet_1.title = "考勤报表"

    try:
        if isinstance(lst, list):
            for row in lst:
                sheet_1.append(row)
        else:
            return None
    except Exception as ex:
        print('Error:', ex)
    finally:
        wb.save(filename)
        wb.close()


def list2excel_test(filename, *lst):
    wb = openpyxl.workbook.Workbook()
    try:
        for l in lst:
            temp_sheet = wb.create_sheet()
            for i in l:
                temp_sheet.append(i)
    except Exception as ex:
        print('Error:', ex)
    finally:
        wb.remove(wb.active)
        wb.save(filename)
        wb.close()


digts = {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}


def char2num(c):
    return digts.get(c)


def str2num(s):
    """字符串转数字"""
    return reduce(lambda x, y: x*10+y, map(char2num, s))


def is_num(s):
    digts['.'] = '.'
    if isinstance(s, str):
        for i in s:
            if i in digts:
                pass
            else:
                return False
        return True
    elif isinstance(s, int) or isinstance(s, float):
        return True
    else:
        return False


# 精确时间到0.5小时
def aj_halfhour(flt):
    Balfour = int(flt)
    dec = flt - Balfour
    if 0.25 > dec >= 0:
        Balfour += 0.0
    elif 0.75 > dec >= 0.25:
        Balfour += 0.5
    else:
        Balfour += 1.0
    return Balfour